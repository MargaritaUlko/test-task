from django.core.management.base import BaseCommand
from news.models import Landmark
from openpyxl import load_workbook
import os
from django.contrib.gis.geos import Point

class Command(BaseCommand):
    help = 'Импортирует данные о примечательных местах из XLSX файла'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Путь к файлу XLSX')

    def handle(self, *args, **kwargs):
        file_path = kwargs['file_path']

        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f'Файл не найден: {file_path}'))
            return

        wb = load_workbook(file_path)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, lat, lon, rating = row

            # Приводим широту и долготу к числовому типу, если возможно
            try:
                lat = float(lat) if lat not in [None, ''] else None
                lon = float(lon) if lon not in [None, ''] else None
                rating = float(rating) if rating not in [None, ''] else None
            except ValueError:
                self.stdout.write(self.style.WARNING(f'Ошибка в данных для места: {name}, строка пропущена.'))
                continue

            # Проверка на валидные данные
            if lat is not None and lon is not None and isinstance(rating, (int, float)):
                Landmark.objects.create(name=name, location=Point(lon, lat), rating=rating)

                self.stdout.write(self.style.SUCCESS(f'Импортировано место: {name}'))
            else:
                self.stdout.write(self.style.WARNING(f'Ошибка в данных для места: {name}, строка пропущена.'))

        self.stdout.write(self.style.SUCCESS('Импорт завершен.'))
